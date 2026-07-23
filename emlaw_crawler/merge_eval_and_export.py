"""Gộp kết quả đánh giá 4 tiêu chí (từ các batch JSON) vào emlaw_qa.json,
rồi xuất ra emlaw_qa_evaluated.xlsx.
"""
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).parent
QA_PATH = BASE_DIR / "emlaw_qa.json"
EVAL_BATCH_DIR = Path(
    "/private/tmp/claude-501/-Users-tranhuuduc-Documents-dev-Rag-Graph-LawLaw"
    "/7721b141-7323-40f2-a20b-5bdd90af496a/scratchpad"
)
BATCH_FILES = [f"qa_eval_batch_{i}.json" for i in range(1, 7)]

OUT_JSON_PATH = BASE_DIR / "emlaw_qa_evaluated.json"
OUT_XLSX_PATH = BASE_DIR / "emlaw_qa_evaluated.xlsx"

CRITERIA = [
    "xac_dinh_dung_van_de",
    "trich_dung_dieu_luat",
    "tra_loi_dung_cau_hoi",
    "cau_tra_loi_ro_rang",
]
COLUMNS = ["stt", "question", "answer"] + CRITERIA + ["ghi_chu_dieu_luat"]
HEADER_VI = {
    "stt": "STT",
    "question": "Câu hỏi",
    "answer": "Câu trả lời",
    "xac_dinh_dung_van_de": "1. Xác định đúng vấn đề",
    "trich_dung_dieu_luat": "2. Trích đúng điều luật",
    "tra_loi_dung_cau_hoi": "3. Trả lời đúng câu hỏi",
    "cau_tra_loi_ro_rang": "4. Câu trả lời rõ ràng",
    "ghi_chu_dieu_luat": "Ghi chú verify điều luật",
}
COLUMN_WIDTHS = {
    "stt": 6,
    "question": 45,
    "answer": 90,
    "xac_dinh_dung_van_de": 14,
    "trich_dung_dieu_luat": 14,
    "tra_loi_dung_cau_hoi": 14,
    "cau_tra_loi_ro_rang": 14,
    "ghi_chu_dieu_luat": 45,
}

TRUE_FILL = PatternFill("solid", fgColor="C6EFCE")
FALSE_FILL = PatternFill("solid", fgColor="FFC7CE")


def load_eval_map() -> dict[int, dict]:
    eval_map: dict[int, dict] = {}
    for name in BATCH_FILES:
        path = EVAL_BATCH_DIR / name
        if not path.exists():
            raise FileNotFoundError(f"Thiếu file batch: {path}")
        with open(path, encoding="utf-8") as f:
            batch = json.load(f)
        for item in batch:
            eval_map[item["stt"]] = item
    return eval_map


def main() -> None:
    with open(QA_PATH, encoding="utf-8") as f:
        rows = json.load(f)

    eval_map = load_eval_map()

    missing = [r["stt"] for r in rows if r["stt"] not in eval_map]
    if missing:
        raise ValueError(f"Thiếu đánh giá cho stt: {missing}")

    merged = []
    for row in rows:
        ev = eval_map[row["stt"]]
        merged_row = dict(row)
        for key in CRITERIA + ["ghi_chu_dieu_luat"]:
            merged_row[key] = ev.get(key)
        merged.append(merged_row)

    with open(OUT_JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(merged, f, ensure_ascii=False, indent=2)
    print(f"Đã ghi {len(merged)} dòng vào {OUT_JSON_PATH}")

    wb = Workbook()
    ws = wb.active
    ws.title = "QA_Evaluated"

    ws.append([HEADER_VI[c] for c in COLUMNS])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    wrap_top = Alignment(wrap_text=True, vertical="top")
    wrap_center = Alignment(wrap_text=True, vertical="top", horizontal="center")

    for row in merged:
        ws.append([row.get(col, "") for col in COLUMNS])

    for row_cells in ws.iter_rows(min_row=2):
        for cell, col in zip(row_cells, COLUMNS):
            if col in CRITERIA:
                cell.alignment = wrap_center
                if cell.value is True:
                    cell.value = "TRUE"
                    cell.fill = TRUE_FILL
                elif cell.value is False:
                    cell.value = "FALSE"
                    cell.fill = FALSE_FILL
            else:
                cell.alignment = wrap_top

    for idx, col in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = COLUMN_WIDTHS[col]

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(merged) + 1}"

    wb.save(OUT_XLSX_PATH)
    print(f"Đã ghi excel: {OUT_XLSX_PATH}")

    # Thống kê nhanh
    for crit in CRITERIA:
        n_false = sum(1 for r in merged if r[crit] is False)
        print(f"  {crit}: {n_false}/{len(merged)} bị đánh giá FALSE")


if __name__ == "__main__":
    main()
