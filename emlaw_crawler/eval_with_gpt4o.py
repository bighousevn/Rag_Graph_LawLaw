#!/usr/bin/env python3
"""Script đánh giá câu trả lời pháp luật của mô hình bằng GPT-4o (OpenAI API).

Tập trung gắt gao vào Mức 2 (trich_dung_dieu_luat):
- Đánh gia FALSE nếu điều luật không tồn tại hoặc trích sai.
- Đánh giá FALSE nếu văn bản quy phạm pháp luật đã hết hiệu lực / bị thay thế (vd: Luật Đất đai 2013 bị thay thế bởi Luật Đất đai 2024).
- ĐÁNH GIÁ FALSE nếu các điều luật được trích dẫn KHÔNG ĐỦ CĂN CỨ để giải quyết hoàn chỉnh câu hỏi người dùng.
"""

import argparse
import concurrent.futures
import json
import os
import sys
from pathlib import Path
from typing import Dict, List, Optional

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

from openai import OpenAI
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field

# Thư mục gốc chứa script
BASE_DIR = Path(__file__).parent.resolve()

DEFAULT_INPUT = BASE_DIR / "answer_dat_dai.json"
DEFAULT_OUT_JSON = BASE_DIR / "answer_dat_dai_evaluated_gpt4o.json"
DEFAULT_OUT_XLSX = BASE_DIR / "answer_dat_dai_evaluated_gpt4o.xlsx"

CRITERIA = [
    "xac_dinh_dung_van_de",
    "trich_dung_dieu_luat",
    "tra_loi_dung_cau_hoi",
    "cau_tra_loi_ro_rang",
]

COLUMNS = ["stt", "question", "answer"] + CRITERIA + ["ghi_chu_dieu_luat"]

HEADER_VI = {
    "stt": "STT",
    "question": "Câu hỏi",
    "answer": "Câu trả lời",
    "xac_dinh_dung_van_de": "1. Xác định đúng vấn đề",
    "trich_dung_dieu_luat": "2. Trích đúng điều luật (Đủ căn cứ)",
    "tra_loi_dung_cau_hoi": "3. Trả lời đúng câu hỏi",
    "cau_tra_loi_ro_rang": "4. Câu trả lời rõ ràng",
    "ghi_chu_dieu_luat": "Ghi chú verify điều luật & căn cứ",
}

COLUMN_WIDTHS = {
    "stt": 6,
    "question": 45,
    "answer": 90,
    "xac_dinh_dung_van_de": 14,
    "trich_dung_dieu_luat": 14,
    "tra_loi_dung_cau_hoi": 14,
    "cau_tra_loi_ro_rang": 14,
    "ghi_chu_dieu_luat": 50,
}

TRUE_FILL = PatternFill("solid", fgColor="C6EFCE")
FALSE_FILL = PatternFill("solid", fgColor="FFC7CE")


class EvaluationSchema(BaseModel):
    xac_dinh_dung_van_de: bool = Field(
        description="Answer có xác định đúng bản chất / vấn đề pháp lý cốt lõi mà câu hỏi đặt ra hay không (không lạc đề, không hiểu sai tình huống)."
    )
    trich_dung_dieu_luat: bool = Field(
        description="Tiêu chí Mức 2: Trả về true nếu điều luật tồn tại, nội dung đúng đắn và CÒN HIỆU LỰC hiện hành (vd: Luật Đất đai 2024 thay 2013). Trả về false nếu: 1) Điều luật không tồn tại/trích bịa; 2) Dùng văn bản HẾT HIỆU LỰC/bị thay thế; 3) Không trích điều luật nào hoặc trích lạc đề."
    )
    tra_loi_dung_cau_hoi: bool = Field(
        description="Answer có trả lời trúng và đủ điều người dùng hỏi không (không né tránh, không trả lời chung chung thiếu kết luận)."
    )
    cau_tra_loi_ro_rang: bool = Field(
        description="Answer có trình bày rõ ràng, mạch lạc, dễ hiểu với người không rành luật không."
    )
    ghi_chu_dieu_luat: str = Field(
        description="Ghi rõ các điều luật trích dẫn, tính hiệu lực hiện hành và nhận xét mức độ bao phủ của điều luật so với câu hỏi (chỉ ra phần nào đã có điều luật căn cứ, phần nào chưa)."
    )


SYSTEM_PROMPT = """Bạn là một chuyên gia pháp lý cao cấp và kiểm định viên chất lượng câu trả lời pháp luật tại Việt Nam (đặc biệt trong lĩnh vực Đất đai).
Nhiệm vụ của bạn là đánh giá chất lượng câu trả lời của mô hình AI cho các câu hỏi pháp lý của người dùng theo 4 tiêu chí chuẩn:

1. `xac_dinh_dung_van_de` (boolean):
   Câu trả lời có xác định đúng bản chất / vấn đề pháp lý cốt lõi mà câu hỏi đặt ra hay không (không hiểu sai ý người hỏi, không đi lạc đề).

2. `trich_dung_dieu_luat` (boolean) - QUY TẮC ĐÁNH GIÁ MỨC 2:
   Tiêu chí này kiểm tra Tính tồn tại, Tính đúng đắn và Tính hiệu lực hiện hành của các điều luật được trích dẫn:
   - **BẮT BUỘC ĐÁNH FALSE** nếu vi phạm một trong các điều sau:
     a. **Văn bản hết hiệu lực / bị thay thế**: Tính đến hiện tại, Luật Đất đai 2013 (Luật số 45/2013/QH13) đã BỊ THAY THẾ hoàn toàn bởi Luật Đất đai 2024 (Luật số 31/2024/QH15, hiệu lực từ 01/08/2024). Các Nghị định cũ (43/2014, 01/2017, 148/2020...) đã bị thay thế bởi các Nghị định mới (101/2024, 102/2024, 88/2024...). Nếu câu trả lời căn cứ vào văn bản cũ ĐÃ HẾT HIỆU LỰC -> BẮT BUỘC ĐÁNH FALSE (`trich_dung_dieu_luat = false`).
     b. **Điều luật không tồn tại hoặc trích sai**: Bị bịa số điều/khoản hoặc trích diễn giải sai bản chất nội dung điều luật.
     c. **Không có điều luật nào hoặc trích dẫn lạc đề**: Không đưa ra điều luật nào hoặc trích điều luật hoàn toàn không liên quan đến tình huống.
   - **ĐÁNH GIÁ TRUE**: Nếu các điều luật trích dẫn là TỒN TẠI, CÒN HIỆU LỰC và NỘI DUNG ĐÚNG ĐẮN có liên quan đến vấn đề câu hỏi.
   - **MỨC ĐỘ BAO PHỦ**: Nếu điều luật trích dẫn đúng và còn hiệu lực nhưng CHỈ MỚI BAO PHỦ ĐƯỢC MỘT PHẦN câu hỏi (ví dụ: trích đúng điều luật cấp sổ lần đầu nhưng chưa có điều luật về tranh chấp), Mức 2 VẪN ĐÁNH GIÁ `true`, đồng thời ghi rõ trong `ghi_chu_dieu_luat` phần nào đã bao phủ, phần nào chưa. (Việc trả lời có trọn vẹn 100% câu hỏi hay không sẽ thuộc Tiêu chí 3).

3. `tra_loi_dung_cau_hoi` (boolean):
   Câu trả lời có trả lời trúng và trọn vẹn các thắc mắc người dùng hỏi không (không né tránh, không bỏ dở câu hỏi chính).

4. `cau_tra_loi_ro_rang` (boolean):
   Câu trả lời có được cấu trúc mạch lạc, trình bày rõ ràng, dễ hiểu đối với người dân không chuyên về luật hay không.

5. `ghi_chu_dieu_luat` (string):
   Giải thích ngắn gọn nhưng đầy đủ:
   - Tóm tắt các điều luật được trích dẫn trong câu trả lời.
   - Khẳng định tính hiệu lực (còn hiệu lực hay đã bị thay thế).
   - Ghi rõ điều luật trích dẫn đã trả lời được phần nào / chưa trả lời được phần nào của câu hỏi.
"""


import time


def evaluate_single_qa(
    client: OpenAI, model_name: str, stt: int, question: str, answer: str
) -> dict:
    """Gọi GPT-4o để đánh giá 1 cặp câu hỏi - câu trả lời (có retry khi gặp Rate Limit 429)."""
    user_prompt = f"""[CÂU HỎI CỦA NGƯỜI DÙNG] (STT: {stt}):
{question}

[CÂU TRẢ LỜI CỦA MÔ HÌNH]:
{answer}
"""
    max_retries = 5
    backoff = 4.0

    for attempt in range(1, max_retries + 1):
        try:
            completion = client.beta.chat.completions.parse(
                model=model_name,
                messages=[
                    {"role": "system", "content": SYSTEM_PROMPT},
                    {"role": "user", "content": user_prompt},
                ],
                response_format=EvaluationSchema,
                temperature=0.1,
            )
            parsed: EvaluationSchema = completion.choices[0].message.parsed
            res = parsed.model_dump()
            res["stt"] = stt
            return res
        except Exception as e:
            err_str = str(e)
            if "429" in err_str or "rate_limit" in err_str.lower():
                print(f"⚠️ STT {stt}: Gặp Rate Limit (429), chờ {backoff:.1f}s trước khi thử lại ({attempt}/{max_retries})...")
                time.sleep(backoff)
                backoff *= 1.5
            else:
                if attempt == max_retries:
                    print(f"❌ Lỗi khi đánh giá STT {stt}: {e}", file=sys.stderr)
                    return {
                        "stt": stt,
                        "is_error": True,
                        "xac_dinh_dung_van_de": False,
                        "trich_dung_dieu_luat": False,
                        "tra_loi_dung_cau_hoi": False,
                        "cau_tra_loi_ro_rang": False,
                        "ghi_chu_dieu_luat": f"Lỗi gọi API GPT-4o: {str(e)}",
                    }
                time.sleep(2.0)


def save_to_excel(merged_data: List[dict], out_xlsx_path: Path) -> None:
    """Xuất danh sách kết quả ra file Excel với định dạng chuyên nghiệp."""
    wb = Workbook()
    ws = wb.active
    ws.title = "QA_Evaluated_GPT4o"

    # Header
    ws.append([HEADER_VI[c] for c in COLUMNS])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    wrap_top = Alignment(wrap_text=True, vertical="top")
    wrap_center = Alignment(wrap_text=True, vertical="top", horizontal="center")

    for row in merged_data:
        ws.append([row.get(col, "") for col in COLUMNS])

    for row_cells in ws.iter_rows(min_row=2):
        for cell, col in zip(row_cells, COLUMNS):
            if col in CRITERIA:
                cell.alignment = wrap_center
                if cell.value is True:
                    cell.value = "TRUE"
                    cell.fill = TRUE_FILL
                elif cell.value is False:
                    cell.value = "FALSE"
                    cell.fill = FALSE_FILL
            else:
                cell.alignment = wrap_top

    for idx, col in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = COLUMN_WIDTHS[col]

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = (
        f"A1:{get_column_letter(len(COLUMNS))}{len(merged_data) + 1}"
    )

    wb.save(out_xlsx_path)
    print(f"📊 Đã ghi file Excel: {out_xlsx_path}")


def main():
    parser = argparse.ArgumentParser(
        description="Đánh giá câu trả lời mô hình bằng GPT-4o theo EVAL_GUIDE (tập trung Mức 2)."
    )
    parser.add_argument(
        "--input",
        type=str,
        default=str(DEFAULT_INPUT),
        help="Đường dẫn file JSON đầu vào chứa stt, question, answer.",
    )
    parser.add_argument(
        "--output-json",
        type=str,
        default=str(DEFAULT_OUT_JSON),
        help="Đường dẫn file JSON lưu kết quả đánh giá.",
    )
    parser.add_argument(
        "--output-xlsx",
        type=str,
        default=str(DEFAULT_OUT_XLSX),
        help="Đường dẫn file Excel lưu kết quả đánh giá.",
    )
    parser.add_argument(
        "--start",
        type=int,
        default=None,
        help="STT bắt đầu đánh giá (1-indexed, inclusive).",
    )
    parser.add_argument(
        "--end",
        type=int,
        default=None,
        help="STT kết thúc đánh giá (1-indexed, inclusive).",
    )
    parser.add_argument(
        "--model",
        type=str,
        default="gpt-4o",
        help="Tên mô hình OpenAI (mặc định: gpt-4o).",
    )
    parser.add_argument(
        "--api-key",
        type=str,
        default=None,
        help="OpenAI API Key (nếu không truyền sẽ lấy từ env OPENAI_API_KEY).",
    )
    parser.add_argument(
        "--concurrency",
        type=int,
        default=1,
        help="Số lượng luồng gọi API song song (mặc định: 1 - an toàn tối đa).",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Đánh giá lại toàn bộ kể cả các câu đã có trong file output JSON.",
    )

    args = parser.parse_args()

    api_key = args.api_key or os.environ.get("OPENAI_API_KEY")
    if not api_key:
        print(
            "❌ KHÔNG TÌM THẤY OPENAI API KEY! Vui lòng thiết lập biến môi trường OPENAI_API_KEY hoặc truyền qua --api-key.",
            file=sys.stderr,
        )
        sys.exit(1)

    client = OpenAI(api_key=api_key)

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"❌ File đầu vào không tồn tại: {input_path}", file=sys.stderr)
        sys.exit(1)

    with open(input_path, encoding="utf-8") as f:
        qa_data = json.load(f)

    # Lọc phạm vi STT
    if args.start is not None or args.end is not None:
        start_stt = args.start if args.start is not None else 1
        end_stt = args.end if args.end is not None else float("inf")
        target_qa = [r for r in qa_data if start_stt <= r["stt"] <= end_stt]
    else:
        target_qa = qa_data

    out_json_path = Path(args.output_json)
    existing_evals: Dict[int, dict] = {}

    if out_json_path.exists() and not args.force:
        try:
            with open(out_json_path, encoding="utf-8") as f:
                old_list = json.load(f)
                for item in old_list:
                    stt = item.get("stt")
                    if stt is not None and "trich_dung_dieu_luat" in item:
                        note = str(item.get("ghi_chu_dieu_luat", ""))
                        if not item.get("is_error") and not note.startswith("Lỗi gọi API GPT-4o"):
                            existing_evals[stt] = item
            print(f"ℹ️  Đã tìm thấy {len(existing_evals)} câu đã được đánh giá hợp lệ từ trước trong {out_json_path.name}")
        except Exception as e:
            print(f"⚠️  Không thể đọc file output cũ: {e}")

    items_to_eval = [
        r for r in target_qa if r["stt"] not in existing_evals or args.force
    ]
    # Sắp xếp theo STT tăng dần để xử lý tuần tự từ câu nhỏ đến lớn
    items_to_eval.sort(key=lambda x: x["stt"])

    completed_already = len(target_qa) - len(items_to_eval)
    if completed_already > 0 and items_to_eval:
        first_next = items_to_eval[0]["stt"]
        print(f"⏩ Bỏ qua {completed_already} câu đã đánh giá xong. Tiếp tục đánh giá từ câu tiếp theo: STT {first_next}")

    print(
        f"🚀 Tổng câu cần xử lý trong phạm vi: {len(target_qa)}. Đã hoàn thành: {completed_already}. Cần đánh giá mới: {len(items_to_eval)}."
    )

    eval_results: Dict[int, dict] = dict(existing_evals)

    if items_to_eval:
        print(f"⚡ Bắt đầu gọi {args.model} với concurrency = {args.concurrency}...")
        with concurrent.futures.ThreadPoolExecutor(
            max_workers=args.concurrency
        ) as executor:
            future_to_stt = {
                executor.submit(
                    evaluate_single_qa,
                    client,
                    args.model,
                    item["stt"],
                    item["question"],
                    item["answer"],
                ): item["stt"]
                for item in items_to_eval
            }

            completed_count = 0
            for future in concurrent.futures.as_completed(future_to_stt):
                stt = future_to_stt[future]
                completed_count += 1
                try:
                    res = future.result()
                    eval_results[stt] = res
                    status_m2 = (
                        "✅ ĐỦ CĂN CỨ" if res["trich_dung_dieu_luat"] else "❌ KHÔNG ĐỦ / SAI MỨC 2"
                    )
                    print(
                        f"[{completed_count}/{len(items_to_eval)}] STT {stt} -> {status_m2} | Note: {res['ghi_chu_dieu_luat'][:70]}..."
                    )
                except Exception as exc:
                    print(f"❌ STT {stt} sinh ngoại lệ: {exc}", file=sys.stderr)

                # Lưu lũy tiến kết quả ra JSON phòng sự cố ngắt đột ngột
                merged_output = []
                for row in qa_data:
                    stt_val = row["stt"]
                    if stt_val in eval_results:
                        merged_row = dict(row)
                        merged_row.update(eval_results[stt_val])
                        merged_output.append(merged_row)

                out_json_path.parent.mkdir(parents=True, exist_ok=True)
                with open(out_json_path, "w", encoding="utf-8") as f:
                    json.dump(merged_output, f, ensure_ascii=False, indent=2)

    # Đã xong tất cả, tạo file Excel
    merged_output = []
    for row in qa_data:
        stt_val = row["stt"]
        if stt_val in eval_results:
            merged_row = dict(row)
            merged_row.update(eval_results[stt_val])
            merged_output.append(merged_row)

    out_json_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_json_path, "w", encoding="utf-8") as f:
        json.dump(merged_output, f, ensure_ascii=False, indent=2)
    print(f"💾 Đã ghi {len(merged_output)} câu vào JSON: {out_json_path}")

    out_xlsx_path = Path(args.output_xlsx)
    out_xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    save_to_excel(merged_output, out_xlsx_path)

    # In bảng thống kê tỷ lệ ĐÚNG / SAI
    print("\n" + "=" * 50)
    print("📈 BẢNG THỐNG KÊ KẾT QUẢ ĐÁNH GIÁ MÔ HÌNH")
    print("=" * 50)
    total_eval = len(merged_output)
    if total_eval > 0:
        for crit in CRITERIA:
            n_true = sum(1 for r in merged_output if r.get(crit) is True)
            n_false = sum(1 for r in merged_output if r.get(crit) is False)
            pct_true = (n_true / total_eval) * 100
            print(
                f" - {HEADER_VI[crit]}: {n_true}/{total_eval} ĐÚNG ({pct_true:.1f}%) | {n_false} FALSE"
            )
    print("=" * 50)


if __name__ == "__main__":
    main()
