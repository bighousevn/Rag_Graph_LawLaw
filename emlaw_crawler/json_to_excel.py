"""Chuyển emlaw_qa.json sang file Excel (.xlsx)."""
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

INPUT_PATH = Path(__file__).parent / "answer_dat_dai_evaluated_gpt4o.json"
OUTPUT_PATH = Path(__file__).parent / "answer_dat_dai_evaluated_gpt4o.xlsx"

CRITERIA = [
    "xac_dinh_dung_van_de",
    "trich_dung_dieu_luat",
    "tra_loi_dung_cau_hoi",
    "cau_tra_loi_ro_rang",
]

HEADER_VI = {
    "stt": "STT",
    "question": "Câu hỏi",
    "answer": "Câu trả lời",
    "xac_dinh_dung_van_de": "1. Xác định đúng vấn đề",
    "trich_dung_dieu_luat": "2. Trích đúng điều luật (Đủ căn cứ)",
    "tra_loi_dung_cau_hoi": "3. Trả lời đúng câu hỏi",
    "cau_tra_loi_ro_rang": "4. Câu trả lời rõ ràng",
    "ghi_chu_dieu_luat": "Ghi chú verify điều luật & căn cứ",
}

COLUMN_WIDTHS = {
    "stt": 8,
    "question": 45,
    "answer": 90,
    "xac_dinh_dung_van_de": 14,
    "trich_dung_dieu_luat": 14,
    "tra_loi_dung_cau_hoi": 14,
    "cau_tra_loi_ro_rang": 14,
    "ghi_chu_dieu_luat": 50,
}

TRUE_FILL = PatternFill("solid", fgColor="C6EFCE")
FALSE_FILL = PatternFill("solid", fgColor="FFC7CE")


def main() -> None:
    with open(INPUT_PATH, encoding="utf-8") as f:
        rows = json.load(f)

    if not rows:
        print("File JSON trống!")
        return

    # Tự động xác định các cột có trong JSON
    sample_row = rows[0]
    has_eval = any(k in sample_row for k in CRITERIA)

    if has_eval:
        columns = ["stt", "question", "answer"] + CRITERIA + ["ghi_chu_dieu_luat"]
    else:
        columns = ["stt", "question", "answer"]

    wb = Workbook()
    ws = wb.active
    ws.title = "QA_Evaluated"

    ws.append([HEADER_VI.get(c, c) for c in columns])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    wrap_top = Alignment(wrap_text=True, vertical="top")
    wrap_center = Alignment(wrap_text=True, vertical="top", horizontal="center")

    for row in rows:
        ws.append([row.get(col, "") for col in columns])

    for row_cells in ws.iter_rows(min_row=2):
        for cell, col in zip(row_cells, columns):
            if col in CRITERIA:
                cell.alignment = wrap_center
                if cell.value is True:
                    cell.value = "TRUE"
                    cell.fill = TRUE_FILL
                elif cell.value is False:
                    cell.value = "FALSE"
                    cell.fill = FALSE_FILL
            else:
                cell.alignment = wrap_top

    for idx, col in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = COLUMN_WIDTHS.get(col, 15)

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"

    wb.save(OUTPUT_PATH)
    print(f"Đã ghi {len(rows)} dòng vào {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
