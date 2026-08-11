"""
Hỏi chatbot emlaw.vn cho file CauHoi_BoSung_Kind1-4.xlsx: cột A = câu hỏi (từ dòng 2,
dòng 1 là header), cột B = Phân loại (Kind, không đụng tới), cột C = câu trả lời (ghi
trực tiếp vào file này, không tạo file JSON riêng).

Resume tự động: bỏ qua các dòng cột C đã có sẵn nội dung. Có retry khi timeout/rỗng/lỗi.
"""

import os
import time

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

import ask_emlaw
from ask_emlaw import ask_question

# Câu rỗng ở lần chạy trước thường dài hơn trung bình -> chatbot cần lâu hơn 120s mặc định.
ask_emlaw.ANSWER_TIMEOUT_S = 240

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
TARGET_FILE = os.path.join(OUTPUT_DIR, "CauHoi_BoSung_Kind1-4.xlsx")
AUTH_STATE_FILE = os.path.join(OUTPUT_DIR, "emlaw_auth_state.json")

START_ROW = 2       # dòng 1 là header
QUESTION_COL = 1    # cột A
ANSWER_COL = 3      # cột C

HEADLESS = True
MAX_RETRIES = 3
RETRY_DELAY_S = 5.0
DELAY_BETWEEN_QUESTIONS_S = 3.0


def ask_with_retry(page, question: str, max_retries: int = MAX_RETRIES):
    answer, quota_exceeded, timed_out = "", False, True
    for attempt in range(1, max_retries + 1):
        try:
            answer, quota_exceeded, timed_out = ask_question(page, question)
        except Exception as e:
            print(f"      ⚠️  Lỗi lần thử {attempt}/{max_retries}: {e}")
            answer, quota_exceeded, timed_out = "", False, True

        if quota_exceeded:
            return answer, quota_exceeded, timed_out
        if answer and not timed_out:
            return answer, quota_exceeded, timed_out

        if attempt < max_retries:
            print(f"      🔁 {'Timeout' if timed_out else 'Rỗng'}, thử lại ({attempt}/{max_retries})...")
            time.sleep(RETRY_DELAY_S)

    return answer, quota_exceeded, timed_out


def main():
    wb = load_workbook(TARGET_FILE)
    ws = wb.active

    rows_to_do = []
    for r in range(START_ROW, ws.max_row + 1):
        question = ws.cell(row=r, column=QUESTION_COL).value
        existing_answer = ws.cell(row=r, column=ANSWER_COL).value
        if question and str(question).strip() and not (existing_answer and str(existing_answer).strip()):
            rows_to_do.append((r, str(question).strip()))

    total_rows = ws.max_row - START_ROW + 1
    already_done = total_rows - len(rows_to_do)
    print(f"Tổng số câu (dòng {START_ROW}-{ws.max_row}): {total_rows}. Đã có sẵn: {already_done}. Cần hỏi: {len(rows_to_do)}.")

    if not rows_to_do:
        print("Không còn câu nào cần hỏi.")
        return

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=HEADLESS)
        if os.path.isfile(AUTH_STATE_FILE):
            print(f"🔐 Dùng phiên đăng nhập đã lưu: {AUTH_STATE_FILE}")
            context = browser.new_context(storage_state=AUTH_STATE_FILE)
        else:
            print("⚠️  Chưa có phiên đăng nhập, chạy ẩn danh (chạy emlaw_login_setup.py để đăng nhập).")
            context = browser.new_context()
        page = context.new_page()

        for idx, (row, question) in enumerate(rows_to_do, start=1):
            print(f"[{idx}/{len(rows_to_do)}] Dòng {row}: {question[:80]}")

            answer, quota_exceeded, timed_out = ask_with_retry(page, question)

            if quota_exceeded:
                print("⚠️  Đã hết lượt hỏi ẩn danh trong ngày. Dừng script.")
                print(f"   Chạy lại script sau để tiếp tục các câu còn thiếu.")
                break

            ws.cell(row=row, column=ANSWER_COL, value=answer)
            if timed_out:
                print(f"   ⚠️  Vẫn chưa hoàn chỉnh sau {MAX_RETRIES} lần thử ({len(answer)} ký tự).")
            else:
                print(f"   ✅ Đã lưu câu trả lời ({len(answer)} ký tự).")

            wb.save(TARGET_FILE)
            time.sleep(DELAY_BETWEEN_QUESTIONS_S)

        browser.close()

    print(f"\nHoàn tất. Kết quả đã lưu trực tiếp vào: {TARGET_FILE}")


if __name__ == "__main__":
    main()
