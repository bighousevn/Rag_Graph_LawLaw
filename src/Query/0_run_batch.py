"""
Bước 0: Chạy hàng loạt 4 bước (1_extract_triplets → 2_embedding_triplets → 3_query_flow →
4_answer) cho DANH SÁCH câu hỏi đọc từ file Excel, không phải 1 câu hỏi trong input/1_question.txt.

KHÔNG đổi gì trong 4 script — với mỗi câu hỏi: ghi câu hỏi vào input/1_question.txt (đúng file
4 script vẫn đọc/ghi), gọi tuần tự 4 script bằng subprocess (script sau chỉ chạy sau khi script
trước chạy xong). KHÔNG lưu kết quả trung gian của từng câu (mỗi câu chạy xong output/* bị câu
sau ghi đè) — chỉ giữ lại câu hỏi + câu trả lời cuối, ghi TẤT CẢ vào 1 file Excel duy nhất sau khi
xong toàn bộ danh sách.

Input:  input/question.xlsx   cột A, không header, mỗi dòng 1 câu hỏi (bỏ dòng rỗng)
Output: output/batch_result.xlsx   2 cột: Câu hỏi, Câu trả lời — nếu file đã tồn tại thì GHI NỐI
        THÊM vào cuối (không ghi đè), để chạy nhiều đợt --offset/--limit khác nhau mà kết quả
        gộp dần vào 1 file.

Dùng:
    python3 0_run_batch.py                     # chạy hết toàn bộ câu hỏi trong file
    python3 0_run_batch.py --limit 5            # chạy 5 câu đầu (chạy mẫu)
    python3 0_run_batch.py --offset 5 --limit 5 # bỏ qua 5 câu đầu, chạy tiếp 5 câu kế (câu 6-10)

Lưu ý: mỗi câu hỏi chạy lại từ đầu (gọi LLM 2 lần ở bước 1, nạp lại model embedding ở bước 2,
nạp lại index ở bước 3, gọi LLM 1 lần ở bước 4) — không tối ưu tái sử dụng model/index giữa các
câu hỏi. Chấp nhận được ở quy mô vài/vài chục câu hỏi; nếu sau này cần chạy hàng trăm câu hỏi thì
mới đáng để tối ưu (tái sử dụng model đã nạp thay vì gọi subprocess).
"""

import os
import sys
import argparse
import subprocess

import openpyxl

BASE_DIR       = os.path.dirname(os.path.abspath(__file__))
QUESTIONS_XLSX = os.path.join(BASE_DIR, "input", "question.xlsx")
QUESTION_FILE  = os.path.join(BASE_DIR, "input", "1_question.txt")
OUTPUT_DIR     = os.path.join(BASE_DIR, "output")
ANSWER_FILE    = os.path.join(OUTPUT_DIR, "4_answer.txt")
RESULT_XLSX    = os.path.join(OUTPUT_DIR, "batch_result.xlsx")

STEPS = ["1_extract_triplets.py", "2_embedding_triplets.py", "3_query_flow.py", "4_answer.py"]


def log(msg: str) -> None:
    print(msg, flush=True)


def load_questions(offset: int, limit):
    wb = openpyxl.load_workbook(QUESTIONS_XLSX, read_only=True)
    ws = wb.worksheets[0]
    all_questions = []
    for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
        val = row[0]
        if val is not None and str(val).strip():
            all_questions.append(str(val).strip())
    wb.close()

    selected = all_questions[offset:]
    if limit is not None:
        selected = selected[:limit]
    return selected


def run_step(script: str) -> None:
    result = subprocess.run([sys.executable, os.path.join(BASE_DIR, script)], cwd=BASE_DIR)
    if result.returncode != 0:
        raise RuntimeError(f"{script} thất bại (exit code {result.returncode}) — dừng batch.")


def main():
    ap = argparse.ArgumentParser(description="Chạy batch câu hỏi từ file Excel")
    ap.add_argument("--offset", type=int, default=0, help="Bỏ qua N câu đầu (mặc định: 0)")
    ap.add_argument("--limit", type=int, default=None, help="Chỉ chạy N câu kế tiếp (mặc định: hết)")
    args = ap.parse_args()

    questions = load_questions(args.offset, args.limit)
    log(f"Đã tải {len(questions)} câu hỏi từ {QUESTIONS_XLSX} (offset={args.offset})")

    results = []
    for i, question in enumerate(questions, 1):
        log(f"\n{'='*70}\n[{i}/{len(questions)}] {question}\n{'='*70}")

        with open(QUESTION_FILE, "w", encoding="utf-8") as f:
            f.write(question)

        for step in STEPS:
            log(f"\n--- Chạy {step} ---")
            run_step(step)

        answer = ""
        if os.path.exists(ANSWER_FILE):
            with open(ANSWER_FILE, encoding="utf-8") as f:
                answer = f.read().strip()

        results.append((question, answer))

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    if os.path.exists(RESULT_XLSX):
        wb = openpyxl.load_workbook(RESULT_XLSX)
        ws = wb.active
        assert ws is not None
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Kết quả"
        ws.append(["Câu hỏi", "Câu trả lời"])
    for question, answer in results:
        ws.append([question, answer])
    wb.save(RESULT_XLSX)

    log(f"\n\nHoàn tất {len(questions)} câu hỏi. Kết quả → {RESULT_XLSX}")


if __name__ == "__main__":
    main()
